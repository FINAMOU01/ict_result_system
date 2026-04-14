import io
import csv
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse
from django.db import transaction
from django.contrib import messages
from academics.models import Course
from academics.utils import get_grade
from results.models import AnonymousCode, Grade
from academics.views import role_required

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False


@role_required('registra', 'admin')
def decode_results(request, course_id):
    course = get_object_or_404(Course, id=course_id)

    # Only decode if all grades submitted
    total = AnonymousCode.objects.filter(enrollment__course=course).count()
    graded = Grade.objects.filter(
        anonymous_code__enrollment__course=course,
        status=Grade.STATUS_SUBMITTED
    ).count()

    if total == 0:
        messages.error(request, "Aucun étudiant codé pour cette matière.")
        return redirect('course_students', course_id=course_id)

    if graded < total:
        messages.warning(request, f"Attention : seulement {graded}/{total} notes soumises.")

    # Decode: join anonymous codes with student info
    results = AnonymousCode.objects.filter(
        enrollment__course=course
    ).select_related(
        'enrollment__student', 'grade'
    ).order_by('enrollment__student__last_name')

    decoded_data = []
    for anon in results:
        student = anon.enrollment.student
        try:
            grade = anon.grade
            cc = grade.cc_grade
            sn = grade.sn_grade
            final = grade.final_grade
            mention = grade.mention
            status = grade.status
            # Get is_pass based on final score using get_grade function
            _, _, is_pass = get_grade(final) if final is not None else (None, None, None)
        except Grade.DoesNotExist:
            cc = sn = final = None
            mention = '-'
            status = 'non noté'
            is_pass = None

        decoded_data.append({
            'matricule': student.matricule,
            'nom': student.last_name,
            'prenom': student.first_name,
            'code': anon.code,
            'cc': cc,
            'sn': sn,
            'final': final,
            'mention': mention,
            'status': status,
            'is_pass': is_pass,
        })

    # Mark as decoded
    with transaction.atomic():
        AnonymousCode.objects.filter(
            enrollment__course=course,
            status=AnonymousCode.STATUS_GRADED
        ).update(status=AnonymousCode.STATUS_DECODED)

    context = {
        'course': course,
        'decoded_data': decoded_data,
        'total': total,
        'graded': graded,
    }
    return render(request, 'results/decode_results.html', context)


@role_required('registra', 'admin')
def export_excel(request, course_id):
    course = get_object_or_404(Course, id=course_id)

    results = AnonymousCode.objects.filter(
        enrollment__course=course
    ).select_related('enrollment__student', 'grade').order_by('enrollment__student__last_name')

    if HAS_OPENPYXL:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f"{course.code}"

        # Header style
        header_fill = PatternFill(start_color="1B3A6B", end_color="1B3A6B", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        center = Alignment(horizontal='center', vertical='center')

        # Title row
        ws.merge_cells('A1:H1')
        title_cell = ws['A1']
        title_cell.value = f"RÉSULTATS - {course.name} ({course.code})"
        title_cell.font = Font(bold=True, size=13, color="1B3A6B")
        title_cell.alignment = center
        ws.row_dimensions[1].height = 30

        # Subtitle
        ws.merge_cells('A2:H2')
        ws['A2'].value = f"Semestre: {course.semester} | Professeur: {course.professor.get_full_name() if course.professor else 'N/A'}"
        ws['A2'].alignment = center
        ws['A2'].font = Font(italic=True, size=10, color="666666")
        ws.row_dimensions[2].height = 20

        ws.append([])  # blank row

        # Column headers
        headers = ['N°', 'MATRICULE', 'NOM', 'PRÉNOM', 'CODE ANON.', 'CC (30%)', 'SN (70%)', 'NOTE FINALE', 'MENTION', 'RÉSULTAT']
        ws.append(headers)
        header_row = ws.row_dimensions[4]
        header_row.height = 20

        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col_idx)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center

        # Data rows
        pass_fill = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
        fail_fill = PatternFill(start_color="FFEBEE", end_color="FFEBEE", fill_type="solid")

        for i, anon in enumerate(results, 1):
            student = anon.enrollment.student
            try:
                grade = anon.grade
                cc = float(grade.cc_grade) if grade.cc_grade else ''
                sn = float(grade.sn_grade) if grade.sn_grade else ''
                final = float(grade.final_grade) if grade.final_grade else ''
                mention = grade.mention
                result = 'ADMIS' if grade.is_passing else 'RECALÉ'
                row_fill = pass_fill if grade.is_passing else fail_fill
            except Grade.DoesNotExist:
                cc = sn = final = mention = ''
                result = 'N/A'
                row_fill = None

            row_data = [i, student.matricule, student.last_name, student.first_name,
                       anon.code, cc, sn, final, mention, result]
            ws.append(row_data)

            if row_fill:
                for col_idx in range(1, 11):
                    ws.cell(row=4 + i, column=col_idx).fill = row_fill
                    ws.cell(row=4 + i, column=col_idx).alignment = center

        # Column widths
        col_widths = [5, 15, 20, 20, 12, 12, 12, 14, 15, 12]
        for col_idx, width in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = width

        # Stats section
        ws.append([])
        ws.append([])
        ws.append(['STATISTIQUES'])
        ws['A' + str(ws.max_row)].font = Font(bold=True, size=11, color="1B3A6B")

        total = results.count()
        admitted = sum(1 for a in results if hasattr(a, 'grade') and a.grade and a.grade.is_passing)
        ws.append(['Total étudiants', total])
        ws.append(['Admis', admitted])
        ws.append(['Recalés', total - admitted])
        ws.append(['Taux de réussite', f"{round(admitted/total*100, 1) if total else 0}%"])

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="resultats_{course.code}.xlsx"'
        wb.save(response)
        return response

    else:
        # Fallback: CSV
        response = HttpResponse(content_type='text/csv; charset=utf-8-sig')
        response['Content-Disposition'] = f'attachment; filename="resultats_{course.code}.csv"'
        writer = csv.writer(response)
        writer.writerow(['Matricule', 'Nom', 'Prénom', 'Code Anon.', 'CC', 'SN', 'Note Finale', 'Mention'])

        for anon in results:
            student = anon.enrollment.student
            try:
                g = anon.grade
                writer.writerow([student.matricule, student.last_name, student.first_name,
                                  anon.code, g.cc_grade, g.sn_grade, g.final_grade, g.mention])
            except Grade.DoesNotExist:
                writer.writerow([student.matricule, student.last_name, student.first_name,
                                  anon.code, '', '', '', ''])
        return response
