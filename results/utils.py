import io
import csv
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from academics.models import Enrollment, Course
from .models import Grade


def generate_codes_for_course(course):
    """
    Assign sequential anonymous codes (1, 2, 3...) to all enrollments
    in a course. Shuffles by student last name then assigns codes.
    """
    enrollments = list(course.enrollments.select_related('student').order_by('?'))
    for i, enrollment in enumerate(enrollments, start=1):
        enrollment.anonymous_code = i
        enrollment.save(update_fields=['anonymous_code'])

        # Create grade placeholder
        Grade.objects.get_or_create(enrollment=enrollment)

    course.is_coded = True
    from django.utils import timezone
    course.coding_done_at = timezone.now()
    course.save(update_fields=['is_coded', 'coding_done_at'])
    return len(enrollments)


def generate_results_excel(course):
    """
    Generate Excel file with decoded results.
    Columns: Name | ID Number | First Name | Last Name | Email Address | Attendance/10 | Assignment/20 | CA/30 | Final/70 | Course Name | Semester
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"{course.code} Results"

    # Styles
    header_font = Font(name='Calibri', bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill(start_color='1B3A6B', end_color='1B3A6B', fill_type='solid')
    border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    center = Alignment(horizontal='center', vertical='center')
    left_align = Alignment(horizontal='left', vertical='center')
    pass_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
    fail_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')

    # Headers
    headers = ['Name', 'ID Number', 'First Name', 'Last Name', 'Email Address', 'Attendance/10', 'Assignment/20', 'CA/30', 'Final/70', 'Course Name', 'Semester']
    ws.append(headers)
    header_row = 1
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = center

    # Data rows
    enrollments = course.enrollments.select_related('student', 'grade').order_by('student__last_name', 'student__first_name')
    walkin_fill = PatternFill(start_color='FFE8CC', end_color='FFE8CC', fill_type='solid')
    
    for enrollment in enrollments:
        grade = getattr(enrollment, 'grade', None)
        ca = float(grade.cc_score) if grade and grade.cc_score is not None else ''
        final_sn = float(grade.sn_score) if grade and grade.sn_score is not None else ''
        final_score = float(grade.final_score) if grade and grade.final_score is not None else ''
        remarks = 'PASS' if isinstance(final_score, float) and final_score >= 10 else ('FAIL' if isinstance(final_score, float) else 'N/A')

        # Append (Walk-in) note to name if walk-in student
        student_name = enrollment.student.full_name()
        if enrollment.student.is_walkin:
            student_name = f"{student_name} (Walk-in)"

        row = [
            student_name,                                             # Name
            enrollment.student.matricule,                            # ID Number
            enrollment.student.first_name,                           # First Name
            enrollment.student.last_name,                            # Last Name
            enrollment.student.email if enrollment.student.email else '',  # Email Address
            '',                                                      # Attendance/10 (blank)
            '',                                                      # Assignment/20 (blank)
            ca,                                                      # CA/30 (cc_score)
            final_sn,                                                # Final/70 (sn_score)
            f"{course.code} {course.name}",                          # Course Name (full format)
            course.semester.name                                     # Semester
        ]
        ws.append(row)
        current_row = ws.max_row
        for col_idx in range(1, 12):
            cell = ws.cell(row=current_row, column=col_idx)
            cell.border = border
            cell.alignment = left_align if col_idx in [1, 3, 4, 10, 11] else center
            # Apply walk-in fill (orange) for walk-in students, else color pass/fail
            if enrollment.student.is_walkin:
                cell.fill = walkin_fill
            elif remarks == 'PASS':
                cell.fill = pass_fill
            elif remarks == 'FAIL':
                cell.fill = fail_fill

    # Column widths
    col_widths = [20, 15, 15, 15, 18, 12, 14, 10, 10, 35, 15]
    for i, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width

    ws.row_dimensions[header_row].height = 20

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def generate_coding_sheet(course):
    """
    Generate Excel with Code ↔ Matricule mapping (for registra only).
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Coding Sheet (CONFIDENTIAL)"

    header_font = Font(name='Calibri', bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill(start_color='C0392B', end_color='C0392B', fill_type='solid')
    border = Border(left=Side(style='thin'), right=Side(style='thin'),
                    top=Side(style='thin'), bottom=Side(style='thin'))
    center = Alignment(horizontal='center', vertical='center')

    ws.merge_cells('A1:D1')
    ws['A1'] = '⚠ CONFIDENTIAL - REGISTRA USE ONLY ⚠'
    ws['A1'].font = Font(name='Calibri', bold=True, size=14, color='C0392B')
    ws['A1'].alignment = center

    ws.merge_cells('A2:D2')
    ws['A2'] = f"Coding Sheet: {course.code} - {course.name} | {course.semester.name}"
    ws['A2'].font = Font(bold=True, size=11)
    ws['A2'].alignment = center

    ws.append([])
    headers = ['Anonymous Code', 'Matricule', 'Full Name', 'Signature']
    ws.append(headers)
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = center

    enrollments = course.enrollments.select_related('student').order_by('anonymous_code')
    for enrollment in enrollments:
        row = [
            enrollment.anonymous_code,
            enrollment.student.matricule,
            enrollment.student.full_name(),
            ''
        ]
        ws.append(row)
        for col_idx in range(1, 5):
            cell = ws.cell(row=ws.max_row, column=col_idx)
            cell.border = border
            cell.alignment = center
            if col_idx == 3:
                cell.alignment = Alignment(horizontal='left', vertical='center')

    ws.column_dimensions['A'].width = 18
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 30
    ws.column_dimensions['D'].width = 20

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def generate_results_csv(course):
    """
    Generate decoded results as CSV bytes.
    Columns: Name | ID Number | First Name | Last Name | Email Address | 
    Attendance/10 | CC/20 | SN/70 | Final/100 | Letter Grade | Course Name | Semester
    """
    buffer = io.StringIO()
    writer = csv.writer(buffer)
    writer.writerow([
        'Name', 'ID Number', 'First Name', 'Last Name', 'Email Address',
        'Attendance / 10', 'CC / 20', 'SN / 70', 'Final / 100', 'Letter Grade', 'Course Name', 'Semester'
    ])

    enrollments = course.enrollments.select_related('student', 'grade').order_by('student__last_name', 'student__first_name')
    for enrollment in enrollments:
        grade = getattr(enrollment, 'grade', None)
        
        # Extract grade components
        attendance = float(grade.attendance_score) if grade and grade.attendance_score is not None else ''
        cc = float(grade.cc_score) if grade and grade.cc_score is not None else ''
        sn = float(grade.sn_score) if grade and grade.sn_score is not None else ''
        final_score = float(grade.final_score) if grade and grade.final_score is not None else ''
        letter_grade = grade.letter_grade if grade and grade.letter_grade else ''

        writer.writerow([
            enrollment.student.matricule,
            enrollment.student.matricule,
            enrollment.student.first_name,
            enrollment.student.last_name,
            enrollment.student.email if enrollment.student.email else '',
            attendance,
            cc,
            sn,
            final_score,
            letter_grade,
            f"{course.code} {course.name}",
            course.semester.name,
        ])

    return io.BytesIO(buffer.getvalue().encode('utf-8-sig'))
